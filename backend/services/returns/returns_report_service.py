"""Returns report — screen: 1 row = 1 RMZ; export: 1 row = 1 RMZLine."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Any, Literal, Optional

from sqlalchemy import exists, func, or_
from sqlalchemy.orm import Session, aliased

from ...models.customer import Customer
from ...models.order import Order
from ...models.order_item import OrderItem
from ...models.product import Product
from ...models.return_status import ReturnStatus
from ...models.sale_document import SaleDocument
from ...models.stock_document import StockDocument
from ...models.warehouse import Warehouse
from ...models.wms_order_return import WmsOrderReturn
from ...models.wms_refund import WmsRefund
from ...models.wms_rmz_line import RMZLine

DateField = Literal["created", "warehouse_commit", "refund"]
SortField = Literal[
    "date",
    "return_number",
    "order_number",
    "product_lines",
    "qty",
    "accepted",
    "rejected",
    "line_value",
    "status",
]

DECISION_LABELS = {
    "OK": "Przyjęty",
    "DAMAGED": "Uszkodzony",
    "REJECTED": "Odrzucony",
}

EXPORT_HEADERS: list[tuple[str, str]] = [
    ("order_number", "Numer zamówienia"),
    ("return_number", "Numer zwrotu"),
    ("product_name", "Nazwa produktu"),
    ("sku", "SKU"),
    ("ean", "EAN"),
    ("qty_returned", "Ilość"),
    ("qty_accepted", "Ilość przyjęta"),
    ("qty_rejected", "Ilość odrzucona"),
    ("qty_damaged_b", "Uszkodzone B"),
    ("qty_damaged_c", "Uszkodzone C"),
    ("line_value", "Wartość towaru"),
    ("currency", "Waluta"),
    ("exchange_rate", "Kurs"),
    ("purchase_cost_net", "Cena zakupu netto PLN (aktualna)"),
    ("customer_name", "Imię i nazwisko"),
    ("customer_phone", "Numer telefonu"),
    ("customer_email", "E-mail"),
    ("return_date", "Data zwrotu"),
    ("status_name", "Status"),
    ("decision_label", "Decyzja produktowa"),
    ("country", "Kraj"),
    ("source", "Źródło"),
    ("warehouse_name", "Magazyn"),
    ("zpz_number", "Z-PZ"),
    ("correction_number", "Numer korekty"),
    ("warehouse_committed", "Przyjęcie magazynowe"),
]


@dataclass(frozen=True)
class ReturnsReportFilters:
    tenant_id: int
    warehouse_id: Optional[int] = None
    date_from: Optional[datetime] = None
    date_to: Optional[datetime] = None
    date_field: DateField = "created"
    status_id: Optional[int] = None
    decision: Optional[str] = None
    product_query: Optional[str] = None
    order_query: Optional[str] = None
    source: Optional[str] = None
    country: Optional[str] = None
    sort: SortField = "date"
    direction: Literal["asc", "desc"] = "desc"
    page: int = 1
    limit: int = 50


def _customer_display_name(c: Customer | None) -> str:
    if c is None:
        return ""
    parts = [str(c.first_name or "").strip(), str(c.last_name or "").strip()]
    name = " ".join(p for p in parts if p).strip()
    if name:
        return name
    return str(getattr(c, "company_name", None) or "").strip()


def _default_date_from() -> datetime:
    return datetime.utcnow() - timedelta(days=30)


def _line_metrics(line: RMZLine, oi: OrderItem | None) -> tuple[int, int, int, int, int, float]:
    accepted = max(0, int(getattr(line, "accepted_qty", None) or 0))
    rejected = max(0, int(getattr(line, "rejected_qty", None) or 0))
    dmg_b = max(0, int(getattr(line, "damaged_b_qty", None) or 0))
    dmg_c = max(0, int(getattr(line, "damaged_c_qty", None) or 0))
    commercial = accepted + dmg_b + dmg_c
    unit = float(oi.unit_price) if oi is not None and oi.unit_price is not None else 0.0
    return accepted, rejected, dmg_b, dmg_c, commercial, round(commercial * unit, 2)


def _apply_header_filters(q, filters: ReturnsReportFilters, *, zpz: Any, refund: Any):
    """Filters on return/order grain. Product/decision use EXISTS (match parent, display all lines)."""
    q = q.filter(
        WmsOrderReturn.tenant_id == int(filters.tenant_id),
        WmsOrderReturn.deleted_at.is_(None),
    )
    if filters.warehouse_id is not None:
        q = q.filter(WmsOrderReturn.warehouse_id == int(filters.warehouse_id))

    date_from = filters.date_from or _default_date_from()
    date_to = filters.date_to
    df = filters.date_field or "created"
    if df == "warehouse_commit":
        col = zpz.created_at
        q = q.filter(WmsOrderReturn.warehouse_document_id.isnot(None))
    elif df == "refund":
        col = refund.decided_at
        q = q.filter(refund.id.isnot(None))
    else:
        col = WmsOrderReturn.created_at
    q = q.filter(col.isnot(None), col >= date_from)
    if date_to is not None:
        q = q.filter(col <= date_to)

    if filters.status_id is not None:
        q = q.filter(WmsOrderReturn.status_id == int(filters.status_id))

    if filters.decision:
        dec = str(filters.decision).strip().upper()
        match_line = aliased(RMZLine)
        q = q.filter(
            exists().where(
                match_line.rmz_id == WmsOrderReturn.id,
                func.upper(match_line.decision) == dec,
            )
        )

    if filters.product_query:
        pq = f"%{str(filters.product_query).strip()}%"
        match_line = aliased(RMZLine)
        match_prod = aliased(Product)
        q = q.filter(
            exists().where(
                match_line.rmz_id == WmsOrderReturn.id,
                match_prod.id == match_line.product_id,
                or_(
                    match_prod.name.ilike(pq),
                    match_prod.sku.ilike(pq),
                    match_prod.ean.ilike(pq),
                    match_prod.barcode.ilike(pq),
                ),
            )
        )

    if filters.order_query:
        oq = f"%{str(filters.order_query).strip()}%"
        q = q.filter(
            or_(
                Order.number.ilike(oq),
                WmsOrderReturn.rmz_number.ilike(oq),
                Order.external_id.ilike(oq),
            )
        )
    if filters.source:
        q = q.filter(Order.source == str(filters.source).strip())
    if filters.country:
        cq = f"%{str(filters.country).strip()}%"
        q = q.filter(or_(Order.country.ilike(cq), Customer.country_code.ilike(cq)))
    return q


def _returns_base_query(db: Session, filters: ReturnsReportFilters):
    zpz = aliased(StockDocument)
    refund = aliased(WmsRefund)
    q = (
        db.query(WmsOrderReturn, Order, Customer, ReturnStatus, Warehouse, zpz, refund)
        .select_from(WmsOrderReturn)
        .join(Order, Order.id == WmsOrderReturn.order_id)
        .outerjoin(Customer, Customer.id == Order.customer_id)
        .outerjoin(ReturnStatus, ReturnStatus.id == WmsOrderReturn.status_id)
        .outerjoin(Warehouse, Warehouse.id == WmsOrderReturn.warehouse_id)
        .outerjoin(zpz, zpz.id == WmsOrderReturn.warehouse_document_id)
        .outerjoin(refund, refund.rmz_id == WmsOrderReturn.id)
    )
    q = _apply_header_filters(q, filters, zpz=zpz, refund=refund)
    return q, zpz, refund


def _load_corrections_by_return(db: Session, *, tenant_id: int, return_ids: list[int]) -> dict[int, SaleDocument]:
    if not return_ids:
        return {}
    ids_str = [str(i) for i in return_ids]
    rows = (
        db.query(SaleDocument)
        .filter(
            SaleDocument.tenant_id == int(tenant_id),
            SaleDocument.document_kind == "CORRECTION",
            SaleDocument.business_source_type == "RETURN",
            SaleDocument.business_source_id.in_(ids_str),
        )
        .order_by(SaleDocument.created_at.desc(), SaleDocument.id.desc())
        .all()
    )
    out: dict[int, SaleDocument] = {}
    for doc in rows:
        try:
            rid = int(doc.business_source_id)
        except (TypeError, ValueError):
            continue
        if rid not in out:
            out[rid] = doc
    return out


def _line_dict(
    *,
    line: RMZLine,
    oi: OrderItem | None,
    product: Product | None,
    currency: str,
) -> dict[str, Any]:
    accepted, rejected, dmg_b, dmg_c, commercial, line_value = _line_metrics(line, oi)
    decision = str(line.decision or "").strip().upper() or None
    purchase = float(product.purchase_price) if product is not None and product.purchase_price is not None else None
    return {
        "return_line_id": int(line.id),
        "product_id": int(line.product_id) if line.product_id else None,
        "product_name": str(product.name or "") if product else "",
        "sku": str(product.sku or "") if product else "",
        "ean": str(product.ean or product.barcode or "") if product else "",
        "qty_returned": float(line.quantity or 0),
        "qty_accepted": accepted,
        "qty_rejected": rejected,
        "qty_damaged_b": dmg_b,
        "qty_damaged_c": dmg_c,
        "qty_commercial": commercial,
        "decision": decision,
        "decision_label": DECISION_LABELS.get(decision or "", decision or ""),
        "line_value": line_value,
        "currency": currency,
        "purchase_cost_net": purchase,
        "purchase_cost_is_current": True,
    }


def _export_line_dict(
    *,
    line: RMZLine,
    ret: WmsOrderReturn,
    order: Order,
    oi: OrderItem | None,
    product: Product | None,
    customer: Customer | None,
    status: ReturnStatus | None,
    warehouse: Warehouse | None,
    zpz: StockDocument | None,
    refund: WmsRefund | None,
    corr: SaleDocument | None,
) -> dict[str, Any]:
    ld = _line_dict(
        line=line,
        oi=oi,
        product=product,
        currency=str(getattr(order, "currency", None) or "PLN"),
    )
    return {
        **ld,
        "return_id": int(ret.id),
        "return_number": str(ret.rmz_number or ""),
        "order_id": int(order.id),
        "order_number": str(order.number or ""),
        "exchange_rate": None,
        "status_id": int(status.id) if status else None,
        "status_name": str(status.name or "") if status else "",
        "customer_name": _customer_display_name(customer),
        "customer_phone": str(customer.phone or "") if customer else "",
        "customer_email": str(customer.email or "") if customer else "",
        "return_date": ret.created_at.isoformat() if ret.created_at else None,
        "source": str(getattr(order, "source", None) or "") or None,
        "order_channel": str(getattr(order, "order_channel", None) or "") or None,
        "country": str(getattr(order, "country", None) or getattr(customer, "country_code", None) or "") or None,
        "warehouse_id": int(ret.warehouse_id) if ret.warehouse_id else None,
        "warehouse_name": str(warehouse.name or "") if warehouse else "",
        "warehouse_committed": bool(ret.warehouse_document_id),
        "zpz_number": str(zpz.document_number or "") if zpz else None,
        "correction_number": str(corr.document_number or "") if corr else None,
        "correction_issued": corr is not None,
        "refund_amount_header": float(refund.refund_amount) if refund and refund.refund_amount is not None else None,
        "refund_shipping": bool(refund.refund_shipping) if refund else False,
    }


def _products_summary_label(lines: list[dict[str, Any]]) -> str:
    n = len(lines)
    if n == 0:
        return "0 produktów"
    if n == 1:
        name = str(lines[0].get("product_name") or "").strip() or "1 produkt"
        return name
    if n == 2:
        name = str(lines[0].get("product_name") or "").strip() or "produkt"
        return f"{name} + 1"
    return f"{n} produktów"


def _load_lines_for_returns(
    db: Session, *, return_ids: list[int]
) -> dict[int, list[tuple[RMZLine, OrderItem | None, Product | None]]]:
    if not return_ids:
        return {}
    rows = (
        db.query(RMZLine, OrderItem, Product)
        .outerjoin(OrderItem, OrderItem.id == RMZLine.order_item_id)
        .outerjoin(Product, Product.id == RMZLine.product_id)
        .filter(RMZLine.rmz_id.in_(return_ids))
        .order_by(RMZLine.rmz_id.asc(), RMZLine.id.asc())
        .all()
    )
    out: dict[int, list[tuple[RMZLine, OrderItem | None, Product | None]]] = {rid: [] for rid in return_ids}
    for line, oi, product in rows:
        out.setdefault(int(line.rmz_id), []).append((line, oi, product))
    return out


def query_returns_report(db: Session, filters: ReturnsReportFilters) -> dict[str, Any]:
    """Screen read model: 1 item = 1 return with embedded lines + aggregates."""
    q, _zpz, _refund = _returns_base_query(db, filters)

    total_returns = (
        q.with_entities(func.count(func.distinct(WmsOrderReturn.id))).scalar() or 0
    )

    sort = filters.sort or "date"
    # Aggregate sorts require subquery; for common header sorts use columns.
    if sort == "return_number":
        sort_col = WmsOrderReturn.rmz_number
    elif sort == "order_number":
        sort_col = Order.number
    elif sort == "status":
        sort_col = ReturnStatus.name
    else:
        sort_col = WmsOrderReturn.created_at

    if filters.direction == "asc":
        q = q.order_by(sort_col.asc().nullslast(), WmsOrderReturn.id.asc())
    else:
        q = q.order_by(sort_col.desc().nullslast(), WmsOrderReturn.id.desc())

    # Aggregate-based sorts applied in Python on current page is wrong for global sort.
    # For qty/value/product_lines/accepted/rejected: order by created_at then re-sort page only
    # is insufficient. Compute aggregates via correlated subquery for those sorts.
    if sort in ("product_lines", "qty", "accepted", "rejected", "line_value"):
        line_agg = (
            db.query(
                RMZLine.rmz_id.label("rmz_id"),
                func.count(RMZLine.id).label("product_lines"),
                func.coalesce(func.sum(func.coalesce(RMZLine.accepted_qty, 0)), 0).label("accepted_sum"),
                func.coalesce(func.sum(func.coalesce(RMZLine.rejected_qty, 0)), 0).label("rejected_sum"),
                func.coalesce(
                    func.sum(
                        func.coalesce(RMZLine.accepted_qty, 0)
                        + func.coalesce(RMZLine.damaged_b_qty, 0)
                        + func.coalesce(RMZLine.damaged_c_qty, 0)
                    ),
                    0,
                ).label("commercial_sum"),
            )
            .group_by(RMZLine.rmz_id)
            .subquery()
        )
        # Re-build with join for aggregate sort
        q, zpz, refund = _returns_base_query(db, filters)
        q = q.outerjoin(line_agg, line_agg.c.rmz_id == WmsOrderReturn.id)
        agg_map = {
            "product_lines": line_agg.c.product_lines,
            "qty": line_agg.c.commercial_sum,
            "accepted": line_agg.c.accepted_sum,
            "rejected": line_agg.c.rejected_sum,
            "line_value": line_agg.c.commercial_sum,  # value approximated by qty order; exact needs unit join
        }
        sort_col = agg_map[sort]
        if filters.direction == "asc":
            q = q.order_by(sort_col.asc().nullslast(), WmsOrderReturn.id.asc())
        else:
            q = q.order_by(sort_col.desc().nullslast(), WmsOrderReturn.id.desc())

    page = max(1, int(filters.page or 1))
    limit = min(100, max(1, int(filters.limit or 50)))
    offset = (page - 1) * limit
    headers = q.offset(offset).limit(limit).all()

    return_ids = [int(ret.id) for ret, *_rest in headers]
    corr_by = _load_corrections_by_return(db, tenant_id=filters.tenant_id, return_ids=return_ids)
    lines_by = _load_lines_for_returns(db, return_ids=return_ids)

    items: list[dict[str, Any]] = []
    for ret, order, customer, status, warehouse, zpz_row, refund_row in headers:
        currency = str(getattr(order, "currency", None) or "PLN")
        raw_lines = lines_by.get(int(ret.id), [])
        line_dicts = [
            _line_dict(line=ln, oi=oi, product=prod, currency=currency) for ln, oi, prod in raw_lines
        ]
        accepted_sum = sum(int(x["qty_accepted"]) for x in line_dicts)
        rejected_sum = sum(int(x["qty_rejected"]) for x in line_dicts)
        commercial_sum = sum(int(x["qty_commercial"]) for x in line_dicts)
        value_sum = round(sum(float(x["line_value"]) for x in line_dicts), 2)
        dmg_b_sum = sum(int(x["qty_damaged_b"]) for x in line_dicts)
        dmg_c_sum = sum(int(x["qty_damaged_c"]) for x in line_dicts)
        corr = corr_by.get(int(ret.id))
        items.append(
            {
                "return": {
                    "return_id": int(ret.id),
                    "return_number": str(ret.rmz_number or ""),
                    "order_id": int(order.id),
                    "order_number": str(order.number or ""),
                    "return_date": ret.created_at.isoformat() if ret.created_at else None,
                    "status_id": int(status.id) if status else None,
                    "status_name": str(status.name or "") if status else "",
                    "customer_name": _customer_display_name(customer),
                    "source": str(getattr(order, "source", None) or "") or None,
                    "country": str(
                        getattr(order, "country", None) or getattr(customer, "country_code", None) or ""
                    )
                    or None,
                    "warehouse_id": int(ret.warehouse_id) if ret.warehouse_id else None,
                    "warehouse_name": str(warehouse.name or "") if warehouse else "",
                    "warehouse_committed": bool(ret.warehouse_document_id),
                    "zpz_number": str(zpz_row.document_number or "") if zpz_row else None,
                    "correction_number": str(corr.document_number or "") if corr else None,
                    "correction_issued": corr is not None,
                    "currency": currency,
                },
                "aggregates": {
                    "product_lines": len(line_dicts),
                    "quantity": commercial_sum,
                    "accepted_qty": accepted_sum,
                    "rejected_qty": rejected_sum,
                    "damaged_b_qty": dmg_b_sum,
                    "damaged_c_qty": dmg_c_sum,
                    "value_gross": value_sum,
                    "products_label": _products_summary_label(line_dicts),
                },
                "lines": line_dicts,
            }
        )

    pages = (int(total_returns) + limit - 1) // limit if limit else 1
    return {
        "items": items,
        "page": page,
        "limit": limit,
        "total": int(total_returns),
        "total_returns": int(total_returns),
        "pages": pages,
    }


def summarize_returns_report(db: Session, filters: ReturnsReportFilters) -> dict[str, Any]:
    """KPI over full filtered set (all matching returns' lines)."""
    q, _zpz, _refund = _returns_base_query(db, filters)
    return_ids = [int(rid) for (rid,) in q.with_entities(WmsOrderReturn.id).distinct().all()]
    if not return_ids:
        return {
            "returns_count": 0,
            "pieces_commercial": 0,
            "value_total": 0.0,
            "accepted_warehouse_qty": 0,
            "rejected_qty": 0,
            "currency": "PLN",
        }

    rows = (
        db.query(RMZLine, OrderItem)
        .outerjoin(OrderItem, OrderItem.id == RMZLine.order_item_id)
        .filter(RMZLine.rmz_id.in_(return_ids))
        .all()
    )
    pieces = 0
    accepted = 0
    rejected = 0
    value = 0.0
    for line, oi in rows:
        a, r, _b, _c, commercial, lv = _line_metrics(line, oi)
        pieces += commercial
        accepted += a
        rejected += r
        value += lv

    return {
        "returns_count": len(return_ids),
        "pieces_commercial": pieces,
        "value_total": round(value, 2),
        "accepted_warehouse_qty": accepted,
        "rejected_qty": rejected,
        "currency": "PLN",
    }


def iter_export_line_rows(db: Session, filters: ReturnsReportFilters, *, batch_returns: int = 100):
    """Export projection: 1 yield = 1 RMZLine across all matching returns."""
    page = 1
    while True:
        chunk = ReturnsReportFilters(
            tenant_id=filters.tenant_id,
            warehouse_id=filters.warehouse_id,
            date_from=filters.date_from,
            date_to=filters.date_to,
            date_field=filters.date_field,
            status_id=filters.status_id,
            decision=filters.decision,
            product_query=filters.product_query,
            order_query=filters.order_query,
            source=filters.source,
            country=filters.country,
            sort=filters.sort,
            direction=filters.direction,
            page=page,
            limit=batch_returns,
        )
        payload = query_returns_report(db, chunk)
        items = payload["items"]
        if not items:
            break
        for group in items:
            ret_meta = group["return"]
            for line in group["lines"]:
                yield {
                    "order_number": ret_meta["order_number"],
                    "return_number": ret_meta["return_number"],
                    "product_name": line["product_name"],
                    "sku": line["sku"],
                    "ean": line["ean"],
                    "qty_returned": line["qty_returned"],
                    "qty_accepted": line["qty_accepted"],
                    "qty_rejected": line["qty_rejected"],
                    "qty_damaged_b": line["qty_damaged_b"],
                    "qty_damaged_c": line["qty_damaged_c"],
                    "line_value": line["line_value"],
                    "currency": line["currency"],
                    "exchange_rate": None,
                    "purchase_cost_net": line.get("purchase_cost_net"),
                    "customer_name": ret_meta.get("customer_name") or "",
                    "customer_phone": "",
                    "customer_email": "",
                    "return_date": ret_meta.get("return_date"),
                    "status_name": ret_meta.get("status_name") or "",
                    "decision_label": line.get("decision_label") or "",
                    "country": ret_meta.get("country"),
                    "source": ret_meta.get("source"),
                    "warehouse_name": ret_meta.get("warehouse_name") or "",
                    "zpz_number": ret_meta.get("zpz_number"),
                    "correction_number": ret_meta.get("correction_number"),
                    "warehouse_committed": ret_meta.get("warehouse_committed"),
                }
        # Enrich phone/email from a dedicated pass would need header customer — already blanked;
        # re-query with full export path for PII columns.
        if page >= int(payload["pages"]):
            break
        page += 1


def iter_export_line_rows_full(db: Session, filters: ReturnsReportFilters, *, batch_returns: int = 80):
    """Export with customer phone/email from header joins (line grain)."""
    page = 1
    while True:
        chunk = ReturnsReportFilters(
            tenant_id=filters.tenant_id,
            warehouse_id=filters.warehouse_id,
            date_from=filters.date_from,
            date_to=filters.date_to,
            date_field=filters.date_field,
            status_id=filters.status_id,
            decision=filters.decision,
            product_query=filters.product_query,
            order_query=filters.order_query,
            source=filters.source,
            country=filters.country,
            sort="date",
            direction="desc",
            page=page,
            limit=batch_returns,
        )
        q, zpz, refund = _returns_base_query(db, chunk)
        q = q.order_by(WmsOrderReturn.created_at.desc().nullslast(), WmsOrderReturn.id.desc())
        offset = (page - 1) * batch_returns
        headers = q.offset(offset).limit(batch_returns).all()
        if not headers:
            break
        return_ids = [int(ret.id) for ret, *_ in headers]
        corr_by = _load_corrections_by_return(db, tenant_id=filters.tenant_id, return_ids=return_ids)
        lines_by = _load_lines_for_returns(db, return_ids=return_ids)
        for ret, order, customer, status, warehouse, zpz_row, refund_row in headers:
            for line, oi, product in lines_by.get(int(ret.id), []):
                yield _export_line_dict(
                    line=line,
                    ret=ret,
                    order=order,
                    oi=oi,
                    product=product,
                    customer=customer,
                    status=status,
                    warehouse=warehouse,
                    zpz=zpz_row,
                    refund=refund_row,
                    corr=corr_by.get(int(ret.id)),
                )
        # Stop when fewer than batch (last page) — also need total awareness
        if len(headers) < batch_returns:
            break
        page += 1


def build_returns_report_csv(db: Session, filters: ReturnsReportFilters) -> bytes:
    buf = io.StringIO()
    buf.write("\ufeff")
    writer = csv.writer(buf, delimiter=";")
    writer.writerow([label for _, label in EXPORT_HEADERS])
    for row in iter_export_line_rows_full(db, filters):
        out = []
        for key, _ in EXPORT_HEADERS:
            val = row.get(key)
            if key == "warehouse_committed":
                out.append("TAK" if val else "NIE")
            elif val is None:
                out.append("")
            else:
                out.append(val)
        writer.writerow(out)
    return buf.getvalue().encode("utf-8")


def build_returns_report_xlsx(db: Session, filters: ReturnsReportFilters) -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Raport zwrotów"
    ws.append([label for _, label in EXPORT_HEADERS])
    for row in iter_export_line_rows_full(db, filters):
        out = []
        for key, _ in EXPORT_HEADERS:
            val = row.get(key)
            if key == "warehouse_committed":
                out.append("TAK" if val else "NIE")
            elif val is None:
                out.append("")
            else:
                out.append(val)
        ws.append(out)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for idx, (_, label) in enumerate(EXPORT_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(label) + 2, 12), 36)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()
